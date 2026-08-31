"""Stage 1 / S1.6 — 湖北省级统计年鉴 xlsx 连接器.

Per docs/20-stage1-s16-provincial-yearbook-plan-20260824.md +
docs/55 §T2 (M1-b, 2026-08-31, knife 627).

Reuses `spikes/02-provincial-yearbook/extract_02_provincial_yearbook.py` for
per-indicator period metadata (B-06), canonical indicator mapping (no Chinese
in DB), and per-row lineage chain (R3-E). Same source of truth used by Stage 0
spike tests.

Single-sample pilot: reads `spikes/02-provincial-yearbook/hubei_2026_06.xlsx`
by default. **No HTTP.** Production multi-period (2020–2025) ingestion is out
of scope for S1.6.

M1 T2 changes (2026-08-31, knife 627):
  * FK resolution lands — hard-coded T1 UUIDs from
    `scripts/seed_m1_reference_data.py` (geo / geo_code_version /
    indicator_definition / indicator_methodology_version / calendar_period).
  * `ingest()` filters observations to only those whose
    `indicator_canonical` has a T1 FK (gdp_cumulative_h1 +
    industrial_value_added_above_threshold); other 19 spike rows are skipped
    to avoid FK failure → PARTIAL.
  * `source_id` reuses T1 `HUBEI_SOURCE_DOC_ID` if a source_document with
    the same SHA already exists (no duplicate doc per SHA).
  * `source_location` row is INSERTed per observation (sheet_name +
    row_locator; composite FK back to source_document).
  * `lineage.source_file_sha256` = file bytes; `caveat_text` non-empty
    (per-row caveat from PERIOD_METADATA_MAP).
  * `comparison_basis` mapped from spike 02's `CUMULATIVE_YOY` /
    `INDEX_YOY` / `PERIOD_END_YOY` into the
    `cegr.comparison_basis` enum (CUMULATIVE / NEEDS_VERIFICATION /
    INSTANTANEOUS).
  * Status semantics tightened: SUCCESS requires ≥1 row inserted. PARTIAL
    or FAILED is not M1 delivery (knife 627 §2.6).

DB ingest flow:
  1. Resolve source_registry row by
     (domain='tjj.hubei.gov.cn', category='PROVINCIAL_BULLETIN').
  2. INSERT ingestion_run (status='RUNNING').
  3. Compute file SHA-256; SELECT-or-INSERT source_document
     (reuse T1 HUBEI_SOURCE_DOC_ID if same SHA exists).
  4. extract() → N observation dicts in memory.
  5. Filter to T1-known indicators (GDP + IAV).
  6. For each filtered observation:
     - INSERT source_location (sheet_name + row_locator; back-link to doc_id).
     - INSERT observation with all T1 FKs + migration-004 columns
       (period_start/end/label/type, lineage JSONB, caveat_text).
  7. UPDATE ingestion_run with final status + counts.

Exit / status semantics:
  SUCCESS — ≥1 filtered observation inserted; ingestion_run row valid.
  PARTIAL — source_document persisted; some obs FK failed (NOT DELIVERY).
  FAILED  — extract failed OR source_document INSERT failed OR no row matched
            T1 FK filter (records_inserted=0).

Red lines (docs/20 §6 + docs/55 §1.2):
  * ❌ 不漂移 CUMULATIVE_HALF_YEAR — period_type stays TEXT; per-indicator
  * ❌ 中文 indicator_zh 不进 DB — only period_label and caveat_text carry
    Chinese; indicator_canonical is snake-case English
  * ❌ 不在 fixture 临时建表 — migration 004 owns schema
  * ❌ 不把首页 HTML 当里程碑（626 CANCELLED）
  * ❌ 不把整表 21 行全插导致未种子指标 FK 失败
  * ❌ 不宣布 Gate / O1 / M1 PASS
"""
from __future__ import annotations

import json
import sys
import uuid
from pathlib import Path

import openpyxl
import psycopg2
import psycopg2.errors

# Path shim so we can import the spike module without copying the parser
REPO_ROOT = Path(__file__).resolve().parents[4]
SPIKE_02_DIR = REPO_ROOT / "spikes" / "02-provincial-yearbook"
if str(SPIKE_02_DIR) not in sys.path:
    sys.path.insert(0, str(SPIKE_02_DIR))

from extract_02_provincial_yearbook import (  # noqa: E402  (sys.path manipulation above)
    INDICATOR_CANONICAL_MAP,
    PROVINCE_CODE_GB2260,
    PROVINCE_ZH,
    SOURCE_AGENCY,
    SOURCE_URL,
    build_lineage_chain,
    compute_sha256 as _spike_compute_sha256,
    extract_rows as _spike_extract_rows,
)


class ProvincialYearbookConnector:
    """Stage 1 / S1.6 + M1 T2 — 湖北省级统计公报 xlsx 连接器."""

    DEFAULT_SAMPLE = SPIKE_02_DIR / "hubei_2026_06.xlsx"
    DEFAULT_REGISTRY_DOMAIN = "tjj.hubei.gov.cn"
    # Administrative category per registry.csv (per Cursor 50 + knife 627 §2.1).
    DEFAULT_REGISTRY_CATEGORY = "PROVINCIAL_BULLETIN"
    DEFAULT_SAMPLE_TITLE = "湖北省2026年1-6月主要经济指标"
    DEFAULT_SAMPLE_PUBLISHER = SOURCE_AGENCY
    DEFAULT_SAMPLE_URL = SOURCE_URL
    DEFAULT_PROVINCE_ZH = PROVINCE_ZH
    DEFAULT_PROVINCE_CODE_GB2260 = PROVINCE_CODE_GB2260
    DEFAULT_EXTRACTOR_VERSION = "2.0"

    # ---- M1 T1 reference UUIDs (must match scripts/seed_m1_reference_data.py) ----
    HUBEI_PROVINCE_ID = uuid.UUID("a1000000-0000-0000-0000-000000000001")
    HUBEI_GEO_CODE_VERSION_ID = uuid.UUID(
        "a1000000-0000-0000-0000-000000000002"
    )
    HUBEI_GDP_INDICATOR_ID = uuid.UUID(
        "a1000000-0000-0000-0000-000000000010"
    )
    HUBEI_GDP_MV_ID = uuid.UUID("a1000000-0000-0000-0000-000000000011")
    HUBEI_IAV_INDICATOR_ID = uuid.UUID(
        "a1000000-0000-0000-0000-000000000020"
    )
    HUBEI_IAV_MV_ID = uuid.UUID("a1000000-0000-0000-0000-000000000021")
    HUBEI_2026_H1_PERIOD_ID = uuid.UUID(
        "a1000000-0000-0000-0000-000020260601"
    )
    HUBEI_SOURCE_DOC_ID = uuid.UUID(
        "a1000000-0000-0000-0000-000000000030"
    )

    # indicator_canonical → (indicator_id, methodology_version_id)
    INDICATOR_FK_MAP: dict[str, tuple[uuid.UUID, uuid.UUID]] = {
        "gdp_cumulative_h1": (HUBEI_GDP_INDICATOR_ID, HUBEI_GDP_MV_ID),
        "industrial_value_added_above_threshold": (
            HUBEI_IAV_INDICATOR_ID,
            HUBEI_IAV_MV_ID,
        ),
    }

    # spike 02 COMPARISON_BASIS_MAP values → cegr.comparison_basis enum.
    # spike 02 uses YoY labels (CUMULATIVE_YOY / INDEX_YOY) which are NOT in
    # the schema enum; map to the closest valid value.
    COMPARISON_BASIS_NORMALIZATION: dict[str, str] = {
        "CUMULATIVE_YOY": "CUMULATIVE",
        "CUMULATIVE_YOY_5MONTH": "CUMULATIVE",
        "PERIOD_END_YOY": "INSTANTANEOUS",
        "INDEX_YOY": "NEEDS_VERIFICATION",
        "UNKNOWN": "NEEDS_VERIFICATION",
    }

    # ---- pure-file helpers ----

    def compute_sha256(self, file_path: Path) -> str:
        """SHA-256 hex digest of `file_path` bytes."""
        if not file_path.exists():
            raise FileNotFoundError(f"Sample xlsx missing: {file_path}")
        return _spike_compute_sha256(file_path)

    def extract(self, file_path: Path) -> dict:
        """Parse xlsx; return dict with sha256 + observations + metadata.

        Pure file operation, no side effects on the DB. Each observation
        carries B-06 period metadata (period_start/end/label/type) and
        `lineage` (chain_id + source_file_sha256 + source_file_url +
        extractor_version) — these are what get written to migration-004
        columns.

        `indicator_zh` is preserved per row for evidence_pack traceability
        but is NEVER written to observation.* columns; only period_label
        and caveat_text may carry Chinese.
        """
        if not file_path.exists():
            raise FileNotFoundError(f"Sample xlsx missing: {file_path}")

        sha = _spike_compute_sha256(file_path)
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        title = None
        for cell in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            title = cell[0]

        col_headers: list[str] = []
        for cell in ws.iter_rows(min_row=2, max_row=2, values_only=True):
            col_headers = [str(c).strip() if c else "" for c in cell]

        rows, footnotes = _spike_extract_rows(ws)

        lineage = build_lineage_chain(sha)
        per_row_lineage = {
            "chain_id": lineage["chain_id"],
            "source_file_sha256": sha,
            "source_file_url": SOURCE_URL,
            "extractor_version": lineage["extractor_version"],
        }
        for r in rows:
            r["lineage"] = per_row_lineage
            # Also store sheet name for source_location row_locator
            r["_sheet_name"] = ws.title

        return {
            "sha256": sha,
            "sheet_name": ws.title,
            "observations": rows,
            "lineage": lineage,
            "metadata": {
                "source_url": SOURCE_URL,
                "province_zh": PROVINCE_ZH,
                "province_code_gb2260": PROVINCE_CODE_GB2260,
                "table_title": str(title).strip() if title else "",
                "column_headers": col_headers,
                "footnote_text": footnotes,
                "extraction_method": (
                    "openpyxl (data_only=True) + spike 02 lineage v2.0"
                ),
                "extractor_version": lineage["extractor_version"],
                "file_name": file_path.name,
                "file_size_bytes": file_path.stat().st_size,
            },
        }

    # ---- DB ingest ----

    def _resolve_source_registry(
        self,
        conn: psycopg2.extensions.connection,
        domain: str = DEFAULT_REGISTRY_DOMAIN,
        category: str = DEFAULT_REGISTRY_CATEGORY,
    ) -> uuid.UUID | None:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT id FROM cegr.source_registry
                WHERE domain = %s AND category = %s
                LIMIT 1
                """,
                (domain, category),
            )
            row = cur.fetchone()
            return row[0] if row else None

    def _create_ingestion_run(
        self,
        conn: psycopg2.extensions.connection,
        source_registry_id: uuid.UUID,
        triggered_by: str,
    ) -> uuid.UUID:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO cegr.ingestion_run
                    (source_registry_id, started_at, status, triggered_by)
                VALUES (%s, NOW(), 'RUNNING', %s)
                RETURNING id
                """,
                (str(source_registry_id), triggered_by),
            )
            run_id = cur.fetchone()[0]
        conn.commit()
        return run_id

    def _resolve_source_document(
        self,
        conn: psycopg2.extensions.connection,
        source_registry_id: uuid.UUID,
        sha256: str,
        size_bytes: int,
        title: str,
        publisher: str,
        url: str,
    ) -> uuid.UUID:
        """Reuse source_document row with same SHA if exists; else INSERT new.

        Per knife 627 §2.4: prefer T1 HUBEI_SOURCE_DOC_ID (same SHA) to
        avoid duplicate source_document per file bytes.
        """
        with conn.cursor() as cur:
            cur.execute(
                "SELECT id FROM cegr.source_document WHERE file_hash_sha256 = %s "
                "ORDER BY created_at ASC LIMIT 1",
                (sha256,),
            )
            row = cur.fetchone()
            if row:
                return row[0]
            cur.execute(
                """
                INSERT INTO cegr.source_document
                    (source_registry_id, source_level, verification_status,
                     title, publisher, url, file_hash_sha256, file_size_bytes,
                     language, extraction_method)
                VALUES (%s, 'S0', 'VERIFIED',
                        %s, %s, %s, %s, %s,
                        'zh', 'EXCEL_PARSE')
                RETURNING id
                """,
                (
                    str(source_registry_id), title, publisher, url,
                    sha256, size_bytes,
                ),
            )
            doc_id = cur.fetchone()[0]
        conn.commit()
        return doc_id

    def _finalize_ingestion_run(
        self,
        conn: psycopg2.extensions.connection,
        run_id: uuid.UUID,
        status: str,
        records_extracted: int,
        records_inserted: int,
        error_log: str | None,
    ) -> None:
        if error_log and len(error_log) > 500:
            error_log = error_log[:497] + "..."
        with conn.cursor() as cur:
            cur.execute(
                """
                UPDATE cegr.ingestion_run
                SET finished_at = NOW(),
                    status = %s,
                    records_extracted = %s,
                    records_inserted = %s,
                    error_log = %s
                WHERE id = %s
                """,
                (status, records_extracted, records_inserted, error_log, str(run_id)),
            )
        conn.commit()

    def _filter_to_known_indicators(
        self, observations: list[dict]
    ) -> list[dict]:
        """Return only observations whose indicator_canonical has a T1 FK."""
        known = set(self.INDICATOR_FK_MAP.keys())
        return [
            obs for obs in observations
            if obs.get("indicator_canonical") in known
        ]

    def _map_comparison_basis(self, basis: str | None) -> str:
        """Map spike 02 COMPARISON_BASIS_MAP values to cegr enum.

        The spike uses YoY labels (CUMULATIVE_YOY, INDEX_YOY, etc.) which
        are NOT in the cegr.comparison_basis enum. Map to the closest
        valid value; unknown → NEEDS_VERIFICATION.
        """
        if not basis:
            return "NEEDS_VERIFICATION"
        return self.COMPARISON_BASIS_NORMALIZATION.get(basis, "NEEDS_VERIFICATION")

    def _insert_observation(
        self,
        conn: psycopg2.extensions.connection,
        observation: dict,
        run_id: uuid.UUID,
        doc_id: uuid.UUID,
        sheet_name: str,
    ) -> tuple[bool, str | None]:
        """Insert one filtered observation: source_location then observation.

        Both rows are written in a single transaction; failure rolls back
        both. Returns (inserted: bool, error_summary: Optional[str]).
        """
        ind_id, mv_id = self.INDICATOR_FK_MAP[observation["indicator_canonical"]]

        lineage = observation.get("lineage") or {}
        lineage_json = json.dumps(lineage, ensure_ascii=False, sort_keys=True)
        # caveat_text must be non-empty per knife 627 §2.5
        caveat = observation.get("caveat") or (
            f"per-row caveat for {observation.get('indicator_canonical')}"
        )
        # unit: spike 02 row carries "亿元" / "%" etc.; may be None for rows
        # where the source did not publish a value (e.g. IAV row in this
        # spike). observation_unit_required CHECK allows unit IS NULL when
        # value IS NULL OR is_imputed = TRUE.
        unit = observation.get("unit") or None

        # missing_reason: required when value IS NULL (per
        # observation_missing_consistency CHECK). For rows where the source
        # did not publish a value we record a per-row provenance note
        # instead of silently dropping the row — the spike 02 parser
        # captures indicator_zh so we can cite the source row.
        if observation.get("value") is None:
            missing_reason = (
                f"value not reported in source row "
                f"{observation.get('row_index', '?')} "
                f"({observation.get('indicator_zh') or observation.get('indicator_canonical')})"
            )
            is_imputed = False
        else:
            missing_reason = None
            is_imputed = False

        try:
            with conn.cursor() as cur:
                # 1. source_location (sheet + row_locator)
                cur.execute(
                    """
                    INSERT INTO cegr.source_location
                        (source_document_id, sheet_name, table_index,
                         row_locator, section_heading)
                    VALUES (%s, %s, %s, %s, %s)
                    RETURNING id
                    """,
                    (
                        str(doc_id),
                        sheet_name,
                        1,
                        f"row {observation.get('row_index', '?')}",
                        observation.get("indicator_zh"),
                    ),
                )
                loc_id = cur.fetchone()[0]

                # 2. observation
                cur.execute(
                    """
                    INSERT INTO cegr.observation
                        (indicator_id, indicator_methodology_version_id,
                         geo_entity_id, geo_code_version_id, calendar_period_id,
                         source_id, source_location_id, ingestion_run_id,
                         value, raw_value, unit, comparison_basis,
                         value_type, status, extraction_method,
                         period_start, period_end, period_label, period_type,
                         lineage, caveat_text, missing_reason, is_imputed)
                    VALUES (
                        %s, %s,
                        %s, %s, %s,
                        %s, %s, %s,
                        %s, %s, %s, %s,
                        'FACT', 'PRELIMINARY', 'EXCEL_PARSE',
                        %s, %s, %s, %s,
                        %s::jsonb, %s, %s, %s
                    )
                    """,
                    (
                        str(ind_id), str(mv_id),
                        str(self.HUBEI_PROVINCE_ID),
                        str(self.HUBEI_GEO_CODE_VERSION_ID),
                        str(self.HUBEI_2026_H1_PERIOD_ID),
                        str(doc_id), str(loc_id), str(run_id),
                        observation.get("value"),
                        str(observation.get("value"))
                        if observation.get("value") is not None else None,
                        unit,
                        self._map_comparison_basis(observation.get("comparison_basis")),
                        observation.get("period_start") or None,
                        observation.get("period_end") or None,
                        observation.get("period_label") or None,
                        observation.get("period_type") or None,
                        lineage_json,
                        caveat,
                        missing_reason,
                        is_imputed,
                    ),
                )
            conn.commit()
            return True, None
        except psycopg2.errors.ForeignKeyViolation as exc:
            conn.rollback()
            return False, f"FK violation: {exc.diag.message_primary or str(exc)[:200]}"
        except psycopg2.errors.UniqueViolation as exc:
            conn.rollback()
            return False, f"UNIQUE violation (natural key dup): {str(exc)[:200]}"
        except psycopg2.errors.Error as exc:
            conn.rollback()
            return False, f"DB error: {str(exc)[:200]}"

    def ingest(
        self,
        file_path: Path,
        conn: psycopg2.extensions.connection,
        triggered_by: str = "provincial_yearbook_connector",
        title: str | None = None,
        publisher: str | None = None,
        url: str | None = None,
    ) -> dict:
        """End-to-end ingest: extract → ingestion_run → source_document →
        observations (T1-FK filtered) → ingestion_run final status.

        Returns a summary dict (suitable for assertions in tests).
        """
        title = title or self.DEFAULT_SAMPLE_TITLE
        publisher = publisher or self.DEFAULT_SAMPLE_PUBLISHER
        url = url or self.DEFAULT_SAMPLE_URL

        sr_id = self._resolve_source_registry(conn)
        if sr_id is None:
            raise RuntimeError(
                "source_registry row for tjj.hubei.gov.cn / PROVINCIAL_BULLETIN "
                "not found; run scripts/import_registry_csv.py first (per M1 T2 · 2026-08-31)"
            )

        run_id = self._create_ingestion_run(conn, sr_id, triggered_by)

        try:
            extracted = self.extract(file_path)
        except FileNotFoundError as exc:
            self._finalize_ingestion_run(
                conn, run_id, "FAILED", 0, 0, f"extract: {exc}"
            )
            return {
                "ingestion_run_id": str(run_id),
                "status": "FAILED",
                "records_extracted": 0,
                "records_inserted": 0,
                "error_log": f"extract: {exc}",
            }

        sha = extracted["sha256"]
        sheet_name = extracted["sheet_name"]
        size_bytes = extracted["metadata"]["file_size_bytes"]
        all_observations = extracted["observations"]

        try:
            doc_id = self._resolve_source_document(
                conn, sr_id, sha, size_bytes, title, publisher, url
            )
        except psycopg2.errors.Error as exc:
            self._finalize_ingestion_run(
                conn, run_id, "FAILED", len(all_observations), 0,
                f"source_document: {str(exc)[:200]}",
            )
            return {
                "ingestion_run_id": str(run_id),
                "status": "FAILED",
                "records_extracted": len(all_observations),
                "records_inserted": 0,
                "error_log": f"source_document: {str(exc)[:200]}",
            }

        filtered = self._filter_to_known_indicators(all_observations)
        n_extracted = len(filtered)
        # records_extracted is what the connector actually attempted to write
        # (= filtered count, not the spike 02 raw row count). This matches
        # the knife §2.6 contract: SUCCESS requires records_inserted≥1.

        inserted = 0
        first_error: str | None = None
        for obs in filtered:
            ok, err = self._insert_observation(conn, obs, run_id, doc_id, sheet_name)
            if ok:
                inserted += 1
            elif first_error is None:
                first_error = err

        # M1-b status semantics (knife 627 §2.6):
        #   0 obs filtered → FAILED (knife requires ≥1 inserted)
        #   all inserted   → SUCCESS
        #   none inserted  → FAILED
        #   partial        → PARTIAL (NOT DELIVERY)
        if n_extracted == 0:
            status = "FAILED"
            first_error = first_error or "no observations matched T1 FK filter"
        elif inserted == n_extracted:
            status = "SUCCESS"
        elif inserted == 0:
            status = "FAILED"
        else:
            status = "PARTIAL"

        self._finalize_ingestion_run(
            conn, run_id, status, n_extracted, inserted, first_error
        )

        return {
            "ingestion_run_id": str(run_id),
            "status": status,
            "records_extracted": n_extracted,
            "records_inserted": inserted,
            "error_log": first_error,
            "source_document_id": str(doc_id),
        }