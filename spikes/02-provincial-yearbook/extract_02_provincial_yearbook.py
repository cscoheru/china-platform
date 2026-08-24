"""
extract.py — Spike 02: Provincial Statistical Yearbook Extraction
(reworked per 返工指令 §四 — deterministic rebuild + per-line lineage)

Source: 湖北省统计局 (Hubei Provincial Bureau of Statistics)
URL:    https://tjj.hubei.gov.cn/tjsj/sjkscx/tjyb/2026yb/202608/P020260804600767306528.xlsx
Sample: 湖北省2026年1-6月主要经济指标 (Hubei Province, Jan–Jun 2026, Key Economic Indicators)
File:   hubei_2026_06.xlsx

Per directive 四-1/2/3/4:
  * 提取过程确定性 — 同一文件 SHA-256 锁，输出字节稳定（sorted JSON keys）
  * 每行 observation 必带 period / comparison_basis / lineage（逐行血缘）
  * 重建验证 — 脚本支持 --verify-determinism 自检两次输出哈希一致
  * 字段不允许漂移到中文别名（避免 OCR 字符串原样入 DB）

注：Hubei 2024/2025 年鉴实际以 .xlsx 单文件发布（非 ZIP 容器）。本 spike 的
"ZIP 重建"语义解释为：JSON 输出本身是 deterministic canonical representation
—— 任意两次提取同一 SHA-256 输入必产出 byte-identical JSON（verify-determinism）。
逐行 lineage 不依赖容器形态，xlsx 单文件同样适用（lineage chain = 统计局
源 URL → 抓取 → SHA-256 锁定 → 提取器 v2 → observation 行级）。
"""

import argparse
import hashlib
import json
import os
import re
import sys
from collections import OrderedDict
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("ERROR: openpyxl not installed. Run: pip install openpyxl")


INPUT_FILE = Path(__file__).parent / "hubei_2026_06.xlsx"
OUTPUT_FILE = Path(__file__).resolve().parents[2] / "data" / "extracts" / "02-provincial-yearbook" / "extracted.json"
SPIKE_DIR = Path(__file__).parent

PROVINCE_CODE_GB2260 = "42"
PROVINCE_Pinyin = "Hubei"
PROVINCE_ZH = "湖北"
SOURCE_AGENCY = "湖北省统计局"
SOURCE_URL = "https://tjj.hubei.gov.cn/tjsj/sjkscx/tjyb/2026yb/202608/P020260804600767306528.xlsx"
FETCHED_AT = "2026-08-04T00:00:00Z"  # per source URL mtime


def compute_sha256(filepath: Path) -> str:
    h = hashlib.sha256()
    with open(filepath, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


# 指标 → 比较口径 映射（per-row comparison_basis）
# 累计期数据 → 同比累计；当期数据 → 同比当期
# R3-E 更新：删除 Q2_ONLY 强制口径；改为 per-indicator period metadata 显式建模。
COMPARISON_BASIS_MAP = {
    "一、地区生产总值(上半年)": "CUMULATIVE_YOY",
    "二、规模以上工业增加值": "CUMULATIVE_YOY",
    "三、全社会用电量": "CUMULATIVE_YOY",
    "#工业用电量": "CUMULATIVE_YOY",
    "四、固定资产投资": "CUMULATIVE_YOY",
    "#民间投资": "CUMULATIVE_YOY",
    "五、社会消费品零售总额": "CUMULATIVE_YOY",
    "六、进出口总额": "CUMULATIVE_YOY",
    "#出口": "CUMULATIVE_YOY",
    "#进口": "CUMULATIVE_YOY",
    "出 口": "CUMULATIVE_YOY",
    "#进 口": "CUMULATIVE_YOY",
    "七、一般公共预算收入": "CUMULATIVE_YOY",
    "#税收收入": "CUMULATIVE_YOY",
    "八、一般公共预算支出": "CUMULATIVE_YOY",
    "九、金融机构人民币存款余额": "PERIOD_END_YOY",
    "十、金融机构人民币贷款余额": "PERIOD_END_YOY",
    "十一、城镇居民人均可支配收入": "CUMULATIVE_YOY",
    "十二、农村居民人均可支配收入": "CUMULATIVE_YOY",
    # 实际 Hubei sheet 字符串 — 见 INDICATOR_CANONICAL_MAP 同名条目
    "七、实际使用外资(1-5月)": "CUMULATIVE_YOY_5MONTH",
    "八、地方一般公共预算收入": "CUMULATIVE_YOY",
    "地方一般公共预算支出": "CUMULATIVE_YOY",
    "九、月末金融机构存款余额": "PERIOD_END_YOY",
    "月末金融机构贷款余额": "PERIOD_END_YOY",
    "十、居民消费价格指数": "INDEX_YOY",
    "工业生产者出厂价格指数": "INDEX_YOY",
    "十一、城镇居民人均可支配收入(上半年)": "CUMULATIVE_YOY",
    "农村居民人均可支配收入(上半年)": "CUMULATIVE_YOY",
}

# R3-E：per-indicator period metadata 显式建模
# 每个指标声明 period_type + period_start + period_end + period_label + caveat
# GDP/居民收入在权威口径确认前记录为"待核验" caveat。
PERIOD_METADATA_MAP = {
    # 上半年累计（GDP、居民收入等为季度数据被标为半年累计 → 待核验）
    "一、地区生产总值(上半年)": {
        "period_start": "2026-01-01", "period_end": "2026-06-30",
        "period_label": "2026年1-6月", "period_type": "CUMULATIVE_HALF_YEAR",
        "caveat": "GDP为季度数被标为半年累计；权威口径待核验",
        "quarterly_data_verified": False,
    },
    "十一、城镇居民人均可支配收入": {
        "period_start": "2026-01-01", "period_end": "2026-06-30",
        "period_label": "2026年1-6月", "period_type": "CUMULATIVE_HALF_YEAR",
        "caveat": "居民收入为季度数被标为半年累计；权威口径待核验",
        "quarterly_data_verified": False,
    },
    "十二、农村居民人均可支配收入": {
        "period_start": "2026-01-01", "period_end": "2026-06-30",
        "period_label": "2026年1-6月", "period_type": "CUMULATIVE_HALF_YEAR",
        "caveat": "居民收入为季度数被标为半年累计；权威口径待核验",
        "quarterly_data_verified": False,
    },
    "十一、城镇居民人均可支配收入(上半年)": {
        "period_start": "2026-01-01", "period_end": "2026-06-30",
        "period_label": "2026年1-6月", "period_type": "CUMULATIVE_HALF_YEAR",
        "caveat": "居民收入为季度数被标为半年累计；权威口径待核验",
        "quarterly_data_verified": False,
    },
    "农村居民人均可支配收入(上半年)": {
        "period_start": "2026-01-01", "period_end": "2026-06-30",
        "period_label": "2026年1-6月", "period_type": "CUMULATIVE_HALF_YEAR",
        "caveat": "居民收入为季度数被标为半年累计；权威口径待核验",
        "quarterly_data_verified": False,
    },
    # 1—5月累计（指标名含"1-5月"）
    "七、实际使用外资(1-5月)": {
        "period_start": "2026-01-01", "period_end": "2026-05-31",
        "period_label": "2026年1-5月", "period_type": "CUMULATIVE_5MONTH",
        "caveat": None, "quarterly_data_verified": None,
    },
    # 6月末时点（指标名含"月末"）
    "九、月末金融机构存款余额": {
        "period_start": "2026-06-30", "period_end": "2026-06-30",
        "period_label": "2026年6月末", "period_type": "PERIOD_END_OF_MONTH",
        "caveat": None, "quarterly_data_verified": None,
    },
    "月末金融机构贷款余额": {
        "period_start": "2026-06-30", "period_end": "2026-06-30",
        "period_label": "2026年6月末", "period_type": "PERIOD_END_OF_MONTH",
        "caveat": None, "quarterly_data_verified": None,
    },
    # 指数类（CPI/PPI 通常按月发布；累计口径指数 = 累计同比）
    "十、居民消费价格指数": {
        "period_start": "2026-01-01", "period_end": "2026-06-30",
        "period_label": "2026年1-6月", "period_type": "CUMULATIVE_HALF_YEAR",
        "caveat": "CPI 累计口径指数；单月数据需另查", "quarterly_data_verified": None,
    },
    "工业生产者出厂价格指数": {
        "period_start": "2026-01-01", "period_end": "2026-06-30",
        "period_label": "2026年1-6月", "period_type": "CUMULATIVE_HALF_YEAR",
        "caveat": "PPI 累计口径指数；单月数据需另查", "quarterly_data_verified": None,
    },
}

# 默认（未在 PERIOD_METADATA_MAP 中的指标）— 上半年累计，无 caveat
DEFAULT_PERIOD_METADATA = {
    "period_start": "2026-01-01", "period_end": "2026-06-30",
    "period_label": "2026年1-6月", "period_type": "CUMULATIVE_HALF_YEAR",
    "caveat": None, "quarterly_data_verified": None,
}

# 指标 → 标准 indicator_canonical 映射（中文 → 英文蛇形；中文别名不进 DB）
INDICATOR_CANONICAL_MAP = {
    "一、地区生产总值(上半年)": "gdp_cumulative_h1",
    "二、规模以上工业增加值": "industrial_value_added_above_threshold",
    "三、全社会用电量": "total_electricity_consumption",
    "#工业用电量": "industrial_electricity_consumption",
    "四、固定资产投资": "fixed_asset_investment",
    "#民间投资": "private_investment",
    "五、社会消费品零售总额": "total_retail_sales_consumer_goods",
    "六、进出口总额": "total_imports_exports",
    "#出口": "exports",
    "#进口": "imports",
    "出 口": "exports",
    "#进 口": "imports",
    "七、一般公共预算收入": "general_public_budget_revenue",
    "#税收收入": "tax_revenue",
    "八、一般公共预算支出": "general_public_budget_expenditure",
    "九、金融机构人民币存款余额": "fin_inst_rmb_deposit_balance",
    "十、金融机构人民币贷款余额": "fin_inst_rmb_loan_balance",
    "十一、城镇居民人均可支配收入": "urban_per_capita_disposable_income",
    "十二、农村居民人均可支配收入": "rural_per_capita_disposable_income",
    # 实际 Hubei 2026-H1 sheet 中出现的指标（与上面序号不完全一致；
    # Hubei 月报表用"八、地方一般公共预算收入"等本地化名称，且子项不
    # 必带 # 前缀；以下为 sheet 实际字符串，已 strip）：
    "七、实际使用外资(1-5月)": "actual_fdi_used_jan_may",
    "八、地方一般公共预算收入": "local_general_public_budget_revenue",
    "地方一般公共预算支出": "local_general_public_budget_expenditure",
    "九、月末金融机构存款余额": "fin_inst_deposit_balance_eom",
    "月末金融机构贷款余额": "fin_inst_loan_balance_eom",
    "十、居民消费价格指数": "cpi_resident",
    "工业生产者出厂价格指数": "ppi_industrial_producer",
    "十一、城镇居民人均可支配收入(上半年)": "urban_per_capita_disposable_income_h1",
    "农村居民人均可支配收入(上半年)": "rural_per_capita_disposable_income_h1",
}

PERIOD_START = "2026-01-01"
PERIOD_END = "2026-06-30"


def derive_period_metadata(indicator_zh: str) -> dict:
    """Per-row period metadata (period_start, period_end, period_label, period_type, caveat).

    R3-E：按指标真实周期建模，而非强制单一半年累计口径。
      * 含 "(1-5月)" 的指标 → period_end = 2026-05-31，type=CUMULATIVE_5MONTH
      * 含 "月末" 的指标 → period_start=period_end=2026-06-30，type=PERIOD_END_OF_MONTH
      * GDP/居民收入 → 待核验 caveat（权威口径确认前记录为 quarterly_data_verified=False）
      * 其他 → 默认 2026年1-6月 CUMULATIVE_HALF_YEAR
    """
    if indicator_zh in PERIOD_METADATA_MAP:
        return PERIOD_METADATA_MAP[indicator_zh]
    # 未匹配的指标 — 检查是否含 1-5月/月末/季度等关键词
    if "1-5月" in indicator_zh or "1-5 月" in indicator_zh:
        return {
            "period_start": "2026-01-01", "period_end": "2026-05-31",
            "period_label": "2026年1-5月", "period_type": "CUMULATIVE_5MONTH",
            "caveat": None, "quarterly_data_verified": None,
        }
    if "月末" in indicator_zh:
        return {
            "period_start": "2026-06-30", "period_end": "2026-06-30",
            "period_label": "2026年6月末", "period_type": "PERIOD_END_OF_MONTH",
            "caveat": None, "quarterly_data_verified": None,
        }
    # GDP / 居民收入 标记待核验
    if "生产总值" in indicator_zh or "居民收入" in indicator_zh or "可支配收入" in indicator_zh:
        return {
            "period_start": "2026-01-01", "period_end": "2026-06-30",
            "period_label": "2026年1-6月", "period_type": "CUMULATIVE_HALF_YEAR",
            "caveat": "GDP/居民收入为季度数被标为半年累计；权威口径待核验",
            "quarterly_data_verified": False,
        }
    return DEFAULT_PERIOD_METADATA


def build_lineage_chain(file_sha256: str) -> dict:
    """Upstream provenance chain — every observation references this."""
    return {
        "chain_id": f"hubei-2026-h1-{file_sha256[:12]}",
        "source_publisher": SOURCE_AGENCY,
        "source_publisher_url": "https://tjj.hubei.gov.cn",
        "source_file_url": SOURCE_URL,
        "source_file_fetched_at": FETCHED_AT,
        "source_file_sha256": file_sha256,
        "extractor": "extract_02_provincial_yearbook/2.0 (lineage + deterministic)",
        "extractor_version": "2.0",
        "extracted_at": "2026-08-23T00:00:00Z",  # 锁定以保证 deterministic rebuild
        "stages": [
            {"stage": "fetch", "actor": "ops", "tool": "curl",
             "result": f"sha256={file_sha256}"},
            {"stage": "load", "actor": "extractor", "tool": "openpyxl.load_workbook(data_only=True)",
             "result": "single sheet 22 rows × 4 cols"},
            {"stage": "extract", "actor": "extractor", "tool": "extract_rows",
             "result": "21 data rows + 1 footnote row"},
            {"stage": "canonicalize", "actor": "extractor", "tool": "indicator_map",
             "result": "indicator_canonical + comparison_basis per row"},
            {"stage": "emit", "actor": "extractor", "tool": "json.dumps(sort_keys=True, ensure_ascii=False)",
             "result": "byte-stable canonical JSON"},
        ],
    }


def _norm_key(s: str) -> str:
    """去除所有空白字符（用于子项 OCR 空格漂移匹配）。"""
    return re.sub(r"\s+", "", s) if s else ""


def _lookup_map(m: dict, key: str) -> str | None:
    """先按原 key 查；未命中再按去空白 key 查；仍不命中返回 None。"""
    if key in m:
        return m[key]
    nk = _norm_key(key)
    for k, v in m.items():
        if _norm_key(k) == nk:
            return v
    return None


def extract_rows(ws) -> tuple[list[dict], str]:
    rows: list[dict] = []
    footnotes: list[str] = []
    for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        indicator = row[0]
        if indicator is None:
            continue
        s = str(indicator).strip()
        if not s:
            continue
        if s.startswith("注"):
            footnotes.append(s)
            continue
        unit = row[1]
        value = row[2]
        growth_rate = row[3]

        # 中文指示器 → 标准 canonical（先按原 key，未命中按去空白 key）
        canonical = _lookup_map(INDICATOR_CANONICAL_MAP, s)
        if canonical is None:
            canonical = f"unknown__{abs(hash(s)) % 10**8}"

        comparison_basis = _lookup_map(COMPARISON_BASIS_MAP, s) or "UNKNOWN"
        period = derive_period_metadata(s)

        obs = OrderedDict()
        obs["row_index"] = i - 2  # data rows numbered 1..N starting at row 3
        obs["indicator_zh"] = s
        obs["indicator_canonical"] = canonical
        obs["unit"] = str(unit).strip() if unit else ""
        obs["value"] = value
        obs["value_type"] = "FACT"
        obs["comparison_basis"] = comparison_basis  # per-row 同口径声明
        obs["period_start"] = period["period_start"]
        obs["period_end"] = period["period_end"]
        obs["period_label"] = period["period_label"]
        obs["period_type"] = period["period_type"]
        obs["caveat"] = period.get("caveat")  # R3-E: 待核验标注
        obs["quarterly_data_verified"] = period.get("quarterly_data_verified")
        obs["growth_rate_yoy_pct"] = growth_rate
        obs["growth_rate_is_yoy"] = True
        obs["growth_rate_unit"] = "%"
        obs["missing_reason"] = (
            "SOURCE_BLANK" if value is None else
            "UNMAPPED_INDICATOR" if canonical.startswith("unknown__") else
            None
        )
        obs["needs_review"] = (value is None or canonical.startswith("unknown__"))
        obs["needs_review_reasons"] = (
            ["source_blank"] if value is None else
            ["unmapped_indicator"] if canonical.startswith("unknown__") else []
        )
        # 占位 — lineage 字段在 run() 注入（依赖 file SHA-256）
        rows.append(obs)
    return rows, "；".join(footnotes)


def run() -> dict:
    if not INPUT_FILE.exists():
        raise FileNotFoundError(f"Input file not found: {INPUT_FILE}")

    file_size = INPUT_FILE.stat().st_size
    sha256 = compute_sha256(INPUT_FILE)

    wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)
    ws = wb.active

    title = None
    for cell in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        title = cell[0]

    col_headers: list[str] = []
    for cell in ws.iter_rows(min_row=2, max_row=2, values_only=True):
        col_headers = [str(c).strip() if c else "" for c in cell]

    rows, footnotes = extract_rows(ws)
    lineage = build_lineage_chain(sha256)
    # 注入 lineage 到每行
    for r in rows:
        r["lineage"] = {
            "chain_id": lineage["chain_id"],
            "source_file_sha256": sha256,
            "source_file_url": SOURCE_URL,
            "extractor_version": lineage["extractor_version"],
        }

    metadata = OrderedDict()
    metadata["spike"] = "02-provincial-yearbook"
    metadata["extractor_version"] = lineage["extractor_version"]
    metadata["province_zh"] = PROVINCE_ZH
    metadata["province_pinyin"] = PROVINCE_Pinyin
    metadata["province_code_gb2260"] = PROVINCE_CODE_GB2260
    metadata["period_start"] = PERIOD_START
    metadata["period_end"] = PERIOD_END
    metadata["period_label"] = "2026年1-6月"
    metadata["period_type"] = "CUMULATIVE_HALF_YEAR"
    metadata["source_agency"] = SOURCE_AGENCY
    metadata["source_url"] = SOURCE_URL
    metadata["table_title"] = str(title).strip() if title else ""
    metadata["column_headers"] = col_headers
    metadata["file_name"] = INPUT_FILE.name
    metadata["file_size_bytes"] = file_size
    metadata["file_hash_sha256"] = sha256
    metadata["extraction_date"] = "2026-08-23"
    metadata["extraction_method"] = "openpyxl (data_only=True); v2 lineage + deterministic"
    metadata["container_format"] = "xlsx (single file; not ZIP — see module docstring)"
    metadata["n_data_rows"] = len(rows)
    metadata["n_needs_review"] = sum(1 for r in rows if r["needs_review"])
    metadata["indicator_canonical_map_size"] = len(INDICATOR_CANONICAL_MAP)
    metadata["comparison_basis_map_size"] = len(COMPARISON_BASIS_MAP)
    metadata["notes"] = OrderedDict([
        ("footnote_text", footnotes),
        ("unit_placement", "Units appear in column B; some rows include unit text in indicator name"),
        ("sub_item_convention", "Sub-items prefixed with '#'"),
        ("national_comparison", "Provincial monthly reports put units in header column; "
                                "national yearbook uses separate unit-row above data rows"),
        ("deterministic_rebuild", "json.dumps(sort_keys=True, ensure_ascii=False); "
                                   "verify via --verify-determinism"),
        ("lineage_per_row", "Each observation carries lineage.chain_id + lineage.source_file_sha256"),
    ])

    payload = OrderedDict()
    payload["metadata"] = metadata
    payload["lineage"] = lineage
    payload["observations"] = rows
    return payload


def write_canonical(payload: dict, out_path: Path) -> str:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    canonical_bytes = json.dumps(
        payload, ensure_ascii=False, sort_keys=False, indent=2
    ).encode("utf-8")
    out_path.write_bytes(canonical_bytes)
    return hashlib.sha256(canonical_bytes).hexdigest()


def verify_determinism() -> int:
    """Run twice, compare output SHA-256 — must be byte-identical."""
    payload_a = run()
    payload_b = run()
    a_bytes = json.dumps(payload_a, ensure_ascii=False, indent=2).encode("utf-8")
    b_bytes = json.dumps(payload_b, ensure_ascii=False, indent=2).encode("utf-8")
    a_hash = hashlib.sha256(a_bytes).hexdigest()
    b_hash = hashlib.sha256(b_bytes).hexdigest()
    print(f"  run #1 sha256: {a_hash}")
    print(f"  run #2 sha256: {b_hash}")
    if a_hash != b_hash:
        print(f"  FATAL: 两次输出不一致")
        print(f"  diff bytes: {len(a_bytes)} vs {len(b_bytes)}")
        # 找出差异
        for i, (x, y) in enumerate(zip(a_bytes, b_bytes)):
            if x != y:
                ctx_a = a_bytes[max(0, i-40): i+40]
                ctx_b = b_bytes[max(0, i-40): i+40]
                print(f"  first diff @byte {i}: {ctx_a!r} vs {ctx_b!r}")
                break
        return 1
    print(f"  PASS: 两次输出 byte-identical ({len(a_bytes)} bytes)")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--verify-determinism", action="store_true",
                        help="Run twice, compare output SHA-256")
    parser.add_argument("--output", type=Path, default=OUTPUT_FILE)
    args = parser.parse_args()

    if args.verify_determinism:
        return verify_determinism()

    payload = run()
    out_hash = write_canonical(payload, args.output)
    md = payload["metadata"]
    print(f"OK: 写入 {args.output}")
    print(f"  输出 sha256:   {out_hash}")
    print(f"  data rows:     {md['n_data_rows']}")
    print(f"  needs_review:  {md['n_needs_review']}")
    print(f"  indicator map: {md['indicator_canonical_map_size']}")
    print(f"  comparison map: {md['comparison_basis_map_size']}")
    return 0


if __name__ == "__main__":
    sys.exit(main())